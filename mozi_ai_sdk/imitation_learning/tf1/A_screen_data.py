import openpyxl
from openpyxl.styles import Font, Alignment

# 数据处理
def main_air():
    sSourceFile = "C:/Users/3-5/Desktop/aircraft_missile/aircraft_zuixin.xlsx"
    sTargetFile = "C:/Users/3-5/Desktop/aircraft_missile/aircraft_2.xlsx"
    wb = openpyxl.load_workbook(sSourceFile)
    worksheet = wb.worksheets
    copy_sheet1 = wb.copy_worksheet(worksheet[0])

    copy_sheet1.title = "Sheet1"

    wb.save(sTargetFile)

    print("It is over")


def delete_worksheet_air():
    work_load = openpyxl.load_workbook(
        "C:/Users/3-5/Desktop/aircraft_missile/aircraft_zuixin.xlsx"
    )
    work_sheets = work_load.worksheets
    list = [88]
    for i in range(500, 1500):
        de_WS = work_sheets[i]
        work_load.remove(de_WS)
        work_load.save("C:/Users/3-5/Desktop/aircraft_missile/aircraft_2.xlsx")


def modify_worksheet_name_air():
    main_air()
    delete_worksheet_air()
    work_load = openpyxl.load_workbook(
        "C:/Users/3-5/Desktop/aircraft_missile/aircraft_2.xlsx"
    )
    worksheet = work_load.worksheets
    num = 0
    for sheet in worksheet:
        num += 1
        num_str = str(num)
        sheet.title = "Sheet" + num_str
    work_load.save("C:/Users/3-5/Desktop/aircraft_missile/aircraft_2.xlsx")


def main_missile():
    sSourceFile = "C:/Users/3-5/Desktop/aircraft_missile/missile_zuixin.xlsx"
    sTargetFile = "C:/Users/3-5/Desktop/aircraft_missile/missile_2.xlsx"
    wb = openpyxl.load_workbook(sSourceFile)
    worksheet = wb.worksheets
    copy_sheet1 = wb.copy_worksheet(worksheet[0])
    # copy_sheet2 = wb.copy_worksheet(wb.worksheets[0])
    # copy_sheet3 = wb.copy_worksheet(wb.worksheets[0])

    copy_sheet1.title = "Sheet1"

    wb.save(sTargetFile)

    print("It is over")


def delete_worksheet_missile():
    work_load = openpyxl.load_workbook(
        "C:/Users/3-5/Desktop/aircraft_missile/missile_zuixin.xlsx"
    )
    work_sheets = work_load.worksheets
    list = [88]
    for i in range(500, 1000):
        de_WS = work_sheets[i]
        work_load.remove(de_WS)
        work_load.save("C:/Users/3-5/Desktop/aircraft_missile/missile_2.xlsx")


def modify_worksheet_name_missile():
    main_missile()
    delete_worksheet_missile()
    work_load = openpyxl.load_workbook(
        "C:/Users/3-5/Desktop/aircraft_missile/missile_2.xlsx"
    )
    worksheet = work_load.worksheets
    num = 0
    for sheet in worksheet:
        num += 1
        num_str = str(num)
        sheet.title = "Sheet" + num_str
    work_load.save("C:/Users/3-5/Desktop/aircraft_missile/missile_2.xlsx")


modify_worksheet_name_missile()
